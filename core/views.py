from django.views.decorators.csrf import csrf_exempt
import csv
import os
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook
from .models import DeliveryRecord
from .forms import DeliveryRecordForm, SearchForm, ExportForm, RestoreForm
from datetime import datetime
import pytz
from django.conf import settings
import zipfile
from io import BytesIO, StringIO
from django.contrib.auth.models import User
from django.utils import timezone
from datetime import timedelta
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from dateutil import parser
from django.core.exceptions import ValidationError
from django.db.models import Count, Avg
from django.db.models.functions import TruncDate
from django.db.models import Count, Avg, F, ExpressionWrapper, fields
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth, ExtractHour, TruncHour
from calendar import monthrange  # Add this import at the top
import xlsxwriter
from io import BytesIO


@login_required
@csrf_exempt
def record_list(request):
    form = SearchForm(request.GET)
    records = DeliveryRecord.objects.all().order_by('serial_number')
    
    if request.user.is_superuser:
        if 'user' in request.GET:
            user_id = request.GET.get('user')
            records = records.filter(created_by__id=user_id)
    # Removed filtering for non-superuser users
    # else:
    #     records = records.filter(created_by=request.user)

    if form.is_valid():
        search_query = form.cleaned_data.get('search_query')
        managed_by = form.cleaned_data.get('managed_by')
        serial_number = form.cleaned_data.get('serial_number')
        min_date = form.cleaned_data.get('min_date')
        max_date = form.cleaned_data.get('max_date')

        if search_query:
            records = records.filter(
                Q(name__icontains=search_query) |
                Q(mrn__icontains=search_query) |
                Q(delivery_date__icontains=search_query) |
                Q(serial_number__icontains=search_query) |
                Q(managed_by__icontains=search_query)
            )

        if managed_by:
            records = records.filter(managed_by__icontains=managed_by)
        if serial_number:
            records = records.filter(serial_number__icontains=serial_number)
        
        if min_date:
            records = records.filter(delivery_date__gte=min_date)
        if max_date:
            records = records.filter(delivery_date__lte=max_date)

    paginator = Paginator(records, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/record_list.html', {'page_obj': page_obj, 'form': form})


@login_required
@csrf_exempt
def record_create(request):
    if request.method == 'POST':
        # Print the POST data for debugging
        print("POST data:", request.POST)
        
        form = DeliveryRecordForm(request.POST)
        if form.is_valid():
            record = form.save(commit=False)
            record.created_by = request.user
            
            # Print the record data before saving
            print("Record delivery_date:", record.delivery_date)
            print("Record delivery_time:", record.delivery_time)
            
            record.save()
            messages.success(request, 'Record successfully added.')
            return redirect('core:record_list')
        else:
            print("Form errors:", form.errors)
            messages.error(request, 'Please correct the errors below.')
    else:
        form = DeliveryRecordForm()
    return render(request, 'core/record_form.html', {'form': form})

@login_required
@csrf_exempt
def record_update(request, pk):
    record = get_object_or_404(DeliveryRecord, pk=pk)
    if request.method == 'POST':
        form = DeliveryRecordForm(request.POST, instance=record)
        if form.is_valid():
            updated_record = form.save(commit=False)
            updated_record.edited_by = request.user
            updated_record.save()
            messages.success(request, 'Record successfully updated.')
            return redirect('core:record_list')
    else:
        # Make sure delivery_date is being passed in the initial data
        form = DeliveryRecordForm(instance=record, initial={
            'delivery_date': record.delivery_date.strftime('%Y-%m-%d') if record.delivery_date else None,
            # ... other initial data ...
        })
    return render(request, 'core/record_form.html', {'form': form})

# @permission_required('core.delete_deliveryrecord')
@csrf_exempt
def record_delete(request, pk):
    record = get_object_or_404(DeliveryRecord, pk=pk)
    if request.method == 'POST':
        record.delete()
        messages.success(request, 'Record successfully deleted.')
        return redirect('core:record_list')
    return render(request, 'core/record_confirm_delete.html', {'record': record})




# @login_required
@csrf_exempt
def export_page(request):
    if request.method == 'POST':
        form = ExportForm(request.POST)
        if form.is_valid():
            min_date = form.cleaned_data['export_min_date']
            max_date = form.cleaned_data['export_max_date']
            
            # Convert dates to timezone-aware datetime objects
            min_datetime = timezone.make_aware(
                datetime.combine(min_date, datetime.min.time())
            )
            max_datetime = timezone.make_aware(
                datetime.combine(max_date, datetime.max.time())
            )

            # Query records with timezone-aware datetimes
            records = DeliveryRecord.objects.filter(
                delivery_date__range=(min_datetime, max_datetime)
            ).order_by('delivery_date')

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Delivery Records"

            # Write headers
            headers = [field.name for field in DeliveryRecord._meta.fields]
            ws.append(headers)

            # Add data rows
            for record in records:
                row = []
                for field in headers:
                    value = getattr(record, field)
                    # Convert User objects to their full name
                    if isinstance(value, User):
                        value = f"{value.first_name} {value.last_name}".strip()  # Use full name
                    # Convert timezone-aware datetimes to naive
                    elif isinstance(value, datetime):
                        value = value.astimezone(pytz.UTC).replace(tzinfo=None)
                    # Convert boolean values to Yes/No
                    elif isinstance(value, bool):
                        value = "Yes" if value else "No"
                    # Handle None/null values
                    elif value is None:
                        value = ""
                    row.append(value)
                ws.append(row)

            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=delivery_records_{min_date}_to_{max_date}.xlsx'
            wb.save(response)
            return response
    else:
        form = ExportForm()
    
    return render(request, 'core/export_page.html', {'form': form})



@csrf_exempt
@login_required
def record_detail(request, pk):
    record = get_object_or_404(DeliveryRecord, pk=pk)
    return render(request, 'core/record_detail.html', {'record': record})



@login_required
@csrf_exempt
def export_single_record(record):
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Record"

    # Write headers
    headers = [field.name for field in DeliveryRecord._meta.fields]
    ws.append(headers)

    # Write data row
    row = []
    for field in headers:
        value = getattr(record, field)
        # Convert User objects to their full name
        if isinstance(value, User):
            value = f"{value.first_name} {value.last_name}".strip()  # Use full name
        # Convert timezone-aware datetimes to naive
        elif isinstance(value, datetime):
            value = value.astimezone(pytz.UTC).replace(tzinfo=None)
        # Convert boolean values to Yes/No
        elif isinstance(value, bool):
            value = "Yes" if value else "No"
        # Handle None/null values
        elif value is None:
            value = ""
        row.append(value)
    ws.append(row)

    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=delivery_record_{record.serial_number}.xlsx'
    wb.save(response)

    return response




@csrf_exempt
@login_required
def backup_records(request):
    try:
        # Get the user's home directory
        home_dir = os.path.expanduser('~')
        # Define the path outside the main directory
        backup_dir = os.path.join(home_dir, 'backup')
        os.makedirs(backup_dir, exist_ok=True)

        # Format timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        csv_filename = f'delivery_records_backup_{timestamp}.csv'
        zip_filename = f'delivery_records_backup_{timestamp}.zip'
        zip_path = os.path.join(backup_dir, zip_filename)

        # Create a ZIP file in memory
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Create CSV content in memory
            csv_content = StringIO()
            writer = csv.writer(csv_content)
            
            # Write header
            headers = [field.name for field in DeliveryRecord._meta.fields]
            writer.writerow(headers)

            # Fetch all records and write data
            records = DeliveryRecord.objects.all().select_related('created_by', 'edited_by').order_by('serial_number')
            for record in records:
                row = []
                for field in headers:
                    value = getattr(record, field)
                    # Convert User objects to their username
                    if field in ['created_by', 'edited_by']:
                        if value:
                            value = value.username  # Store username instead of full name
                    # Convert boolean values to Yes/No
                    elif isinstance(value, bool):
                        value = "Yes" if value else "No"
                    # Handle None/null values
                    elif value is None:
                        value = ""
                    row.append(value)
                writer.writerow(row)

            # Get the CSV content and encode it
            csv_content.seek(0)
            csv_bytes = csv_content.getvalue().encode('utf-8')

            # Add CSV file to ZIP
            zip_file.writestr(f'{csv_filename}', csv_bytes)

        # Save the ZIP file
        with open(zip_path, 'wb') as f:
            f.write(zip_buffer.getvalue())

        # Prepare response
        with open(zip_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'

        messages.success(request, 'Backup folder has been created and is being downloaded.')
        return response

    except Exception as e:
        messages.error(request, f'An error occurred during backup: {str(e)}')
        return redirect('core:record_list')


@user_passes_test(lambda u: u.is_superuser)
def admin_productivity(request):
    # Get time period from request
    time_period = request.GET.get('time_period', 'total')
    
    # Get current date
    today = timezone.now().date()
    
    # Set date range based on time period
    if time_period == 'custom':
        start_date_str = request.GET.get('analytics_start_date')
        end_date_str = request.GET.get('analytics_end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            # Default to last 30 days if dates are invalid
            end_date = today
            start_date = end_date - timedelta(days=30)
    elif time_period == 'total':
        # Use the earliest record date as start date
        first_record = DeliveryRecord.objects.order_by('created_at').first()
        start_date = first_record.created_at.date() if first_record else today
        end_date = today
    else:
        # Default to last 30 days
        end_date = today
        start_date = end_date - timedelta(days=30)

    # Get filter parameters
    date_range = request.GET.get('date_range', 'this_month')
    group_by = request.GET.get('group_by', 'daily')
    custom_start_date = request.GET.get('custom_start_date')
    custom_end_date = request.GET.get('custom_end_date')

    # Calculate date ranges
    today = timezone.now().date()
    
    # Set date range based on selection
    if date_range == 'custom' and custom_start_date and custom_end_date:
        try:
            start_date = datetime.strptime(custom_start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(custom_end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date = today - timedelta(days=30)
            end_date = today
    elif date_range == 'today':
        start_date = end_date = today
    elif date_range == 'yesterday':
        start_date = end_date = today - timedelta(days=1)
    elif date_range == 'this_week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'last_week':
        start_date = today - timedelta(days=today.weekday() + 7)
        end_date = start_date + timedelta(days=6)
    elif date_range == 'this_month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'last_month':
        last_month = today.replace(day=1) - timedelta(days=1)
        start_date = last_month.replace(day=1)
        end_date = last_month
    else:
        start_date = today - timedelta(days=30)
        end_date = today

    # Update context with date range information
    context = {
        'time_period': time_period,
        'analytics_start_date': start_date,
        'analytics_end_date': end_date,
        'date_range': date_range,
        'group_by': group_by,
        'custom_start_date': start_date,
        'custom_end_date': end_date,
    }

    # Get all active users
    users = User.objects.filter(is_active=True)
    
    # Prepare user data
    user_data = []
    total_records = 0
    
    for user in users:
        # Base queryset for user's records within the selected date range
        base_records = DeliveryRecord.objects.filter(
            created_by=user,
            created_at__date__range=[start_date, end_date]
        )
        
        # Calculate counts based on grouping
        if group_by == 'daily':
            period_records = base_records.filter(created_at__date=today)
        elif group_by == 'weekly':
            period_records = base_records.filter(
                created_at__date__range=[today - timedelta(days=7), today]
            )
        else:  # monthly
            period_records = base_records.filter(
                created_at__date__range=[today.replace(day=1), today]
            )
        
        period_count = period_records.count()
        total_count = base_records.count()
        
        # Get last active time
        last_record = DeliveryRecord.objects.filter(created_by=user).order_by('-created_at').first()
        
        if total_count > 0 or user.is_superuser:
            user_data.append({
                'user': user,
                'daily_count': base_records.filter(created_at__date=today).count(),
                'weekly_count': base_records.filter(
                    created_at__date__range=[today - timedelta(days=7), today]
                ).count(),
                'monthly_count': base_records.filter(
                    created_at__date__range=[today.replace(day=1), today]
                ).count(),
                'total_count': total_count,
                'last_active': last_record.created_at if last_record else None
            })
            total_records += total_count

    # Calculate summary statistics
    active_users = sum(1 for data in user_data if data['total_count'] > 0)
    date_diff = (end_date - start_date).days + 1
    avg_daily_records = round(total_records / date_diff if date_diff > 0 else 0)
    
    # Sort user data by total_count in descending order
    user_data.sort(key=lambda x: x['total_count'], reverse=True)

    # Get filter parameters for analytics
    time_period = request.GET.get('time_period', 'daily')  # daily, weekly, monthly
    analysis_range = request.GET.get('analysis_range', '30')  # number of days to analyze
    
    # Calculate date range for analysis
    end_date = timezone.now()
    start_date = end_date - timedelta(days=int(analysis_range))

    # Base queryset for analytics
    base_qs = DeliveryRecord.objects.filter(created_at__range=[start_date, end_date])

    # Top Users Analysis
    top_users = base_qs.values(
        'created_by__username', 
        'created_by__first_name', 
        'created_by__last_name'
    ).annotate(
        record_count=Count('serial_number')
    ).order_by('-record_count')[:5]

    # Average Time to Enter Records
    user_avg_times = []
    all_users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    for user in all_users:
        user_records = base_qs.filter(created_by=user).order_by('created_at')
        record_count = user_records.count()
        
        if record_count > 0:
            if record_count > 1:
                total_time = timedelta()
                count = 0
                prev_record = None
                for record in user_records:
                    if prev_record:
                        diff = record.created_at - prev_record.created_at
                        if diff.total_seconds() < 3600:  # Only count if less than 1 hour difference
                            total_time += diff
                            count += 1
                    prev_record = record
                avg_minutes = (total_time.total_seconds() / count) / 60 if count > 0 else 0
            else:
                avg_minutes = 0
                
            user_avg_times.append({
                'username': user.username,
                'full_name': user.get_full_name() or user.username,
                'avg_minutes': round(avg_minutes, 1),
                'record_count': record_count
            })
    
    # Sort by average time
    user_avg_times.sort(key=lambda x: x['avg_minutes'])

    # Peak Activity Analysis
    trunc_func = {
        'daily': TruncDate,
        'weekly': TruncWeek,
        'monthly': TruncMonth,
        'total': TruncDate,  # Add 'total' option using TruncDate
        'custom': TruncDate  # Add 'custom' option using TruncDate
    }[time_period]

    # If time_period is 'total', adjust the date range to include all records
    if time_period == 'total':
        start_date = DeliveryRecord.objects.order_by('created_at').first().created_at.date() if DeliveryRecord.objects.exists() else today
        end_date = today
    elif time_period == 'custom':
        # Get custom date range from request
        custom_start = request.GET.get('analytics_start_date')
        custom_end = request.GET.get('analytics_end_date')
        if custom_start and custom_end:
            try:
                start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
                end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
            except ValueError:
                start_date = today - timedelta(days=30)
                end_date = today
        else:
            start_date = today - timedelta(days=30)
            end_date = today

    # Update the base_qs to use the correct date range
    base_qs = DeliveryRecord.objects.filter(created_at__date__range=[start_date, end_date])

    peak_activity = base_qs.annotate(
        period=trunc_func('created_at')
    ).values('period').annotate(
        count=Count('serial_number')
    ).order_by('-count')

    # Get the peak period and its details
    if peak_activity:
        peak_period = peak_activity[0]
        peak_users = base_qs.filter(
            created_at__date=peak_period['period']
        ).values(
            'created_by__username',
            'created_by__first_name',
            'created_by__last_name'
        ).annotate(
            count=Count('serial_number')
        ).order_by('-count')
    else:
        peak_period = None
        peak_users = []

    # Calculate statistics for Total Records
    previous_period_records = DeliveryRecord.objects.filter(
        created_at__date__range=[start_date - timedelta(days=date_diff), start_date - timedelta(days=1)]
    ).count()
    
    records_growth = (
        ((total_records - previous_period_records) / previous_period_records * 100)
        if previous_period_records > 0 else 0
    )

    # Calculate statistics for Active Users
    total_users = User.objects.filter(is_active=True).count()
    
    # Calculate statistics for Average Daily Records
    daily_counts = DeliveryRecord.objects.filter(
        created_at__date__range=[start_date, end_date]
    ).annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        count=Count('serial_number')
    ).values_list('count', flat=True)

    highest_daily_records = max(daily_counts) if daily_counts else 0
    lowest_daily_records = min(daily_counts) if daily_counts else 0

    # Calculate statistics for Records This Month
    current_date = timezone.now().date()
    days_in_current_month = monthrange(current_date.year, current_date.month)[1]
    monthly_target = avg_daily_records * days_in_current_month if avg_daily_records else 0
    monthly_target_per_day = round(monthly_target / days_in_current_month) if monthly_target > 0 else 0

    # Convert to JSON-safe format and add to context
    context.update({
        'user_data': user_data,
        'total_records': total_records,
        'active_users': active_users,
        'avg_daily_records': avg_daily_records,
        'records_this_month': DeliveryRecord.objects.filter(
            created_at__date__range=[today.replace(day=1), today]
        ).count(),
        'start_date': start_date,
        'end_date': end_date,
        'top_users': top_users,
        'user_avg_times': user_avg_times,
        'peak_period': peak_period,
        'peak_users': peak_users,
        'time_period': time_period,
        'analysis_range': analysis_range,
        'previous_total_records': previous_period_records,
        'records_growth': round(records_growth, 1),
        'total_users': total_users,
        'highest_daily_records': highest_daily_records,
        'lowest_daily_records': lowest_daily_records,
        'monthly_target': monthly_target,
        'monthly_target_per_day': monthly_target_per_day,
    })
    
    return render(request, 'core/admin_productivity.html', context)

@user_passes_test(lambda u: u.is_superuser)
def export_productivity(request):
    """Export productivity data to Excel"""

    
    # Create an in-memory output file for the new workbook
    output = BytesIO()
    
    # Create the workbook and add a worksheet
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    
    # Add headers
    headers = ['Username', 'Full Name', 'Email', 'Daily Records', 
              'Weekly Records', 'Monthly Records', 'Total Records', 'Last Active']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)
    
    # Get the data (reuse logic from admin_productivity view)
    today = timezone.now().date()
    users = User.objects.filter(is_active=True)
    
    # Write data rows
    for row, user in enumerate(users, start=1):
        user_records = DeliveryRecord.objects.filter(created_by=user)
        daily_count = user_records.filter(created_at__date=today).count()
        weekly_count = user_records.filter(
            created_at__date__range=[today - timedelta(days=7), today]
        ).count()
        monthly_count = user_records.filter(
            created_at__date__range=[today.replace(day=1), today]
        ).count()
        total_count = user_records.count()
        last_record = user_records.order_by('-created_at').first()
        
        data = [
            user.username,
            user.get_full_name(),
            user.email,
            daily_count,
            weekly_count,
            monthly_count,
            total_count,
            last_record.created_at.strftime('%Y-%m-%d %H:%M:%S') if last_record else 'Never'
        ]
        
        for col, value in enumerate(data):
            worksheet.write(row, col, value)
    
    workbook.close()
    
    # Create the HttpResponse object with Excel header
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=productivity_report.xlsx'
    
    return response


# @csrf_exempt
# @login_required
# @user_passes_test(lambda u: u.is_superuser)  # Only allow access to superusers (admin)
# def restore_records(request):
#     if request.method == 'POST':
#         form = RestoreForm(request.POST, request.FILES)
#         if form.is_valid():
#             csv_file = request.FILES['csv_file']
#             if not csv_file.name.endswith('.csv'):
#                 messages.error(request, 'Please upload a CSV file.')
#                 return render(request, 'core/restore_page.html', {'form': form})
            
#             try:
#                 csv_content = csv_file.read().decode('utf-8')
#                 csv_file = StringIO(csv_content)
#                 reader = csv.reader(csv_file)
#                 header = next(reader)
                
#                 for row in reader:
#                     record_data = dict(zip(header, row))
                    
#                     # Check if a record with the same serial_number already exists
#                     if DeliveryRecord.objects.filter(serial_number=record_data.get('serial_number')).exists():
#                         messages.warning(request, f'Record with serial number {record_data.get("serial_number")} already exists. Skipping this record.')
#                         continue  # Skip this record if it already exists
                    
#                     # Convert string values to appropriate types
#                     for key, value in record_data.items():
#                         if value == 'Yes':
#                             record_data[key] = True
#                         elif value == 'No':
#                             record_data[key] = False
#                         elif key == 'delivery_date':
#                             try:
#                                 record_data[key] = parser.parse(value)  # No need for make_aware
#                             except Exception as e:
#                                 messages.error(request, f'Invalid date format in CSV: {value}. Error: {str(e)}')
#                                 return render(request, 'core/restore_page.html', {'form': form})
#                         elif key == 'created_by':  # Assuming this is the column for the user
#                             print(f"Looking up user: '{value}'")  # Debugging line
#                             try:
#                                 # Split the full name into first and last name
#                                 first_name, last_name = value.split(' ', 1)
#                                 user = User.objects.get(first_name=first_name, last_name=last_name)
#                                 record_data[key] = user
#                             except User.DoesNotExist:
#                                 messages.error(request, f'User not found: {value}')
#                                 return render(request, 'core/restore_page.html', {'form': form})
                            
#                     # Create and save the new record
#                     DeliveryRecord.objects.create(**record_data)
                
#                 messages.success(request, 'Records restored successfully.')
#                 return redirect('core:record_list')
            
#             except Exception as e:
#                 messages.error(request, f'An error occurred while processing the CSV file: {str(e)}')
#                 return render(request, 'core/restore_page.html', {'form': form})
    
#     else:
#         form = RestoreForm()
    
#     return render(request, 'core/restore_page.html', {'form': form})


@csrf_exempt
@user_passes_test(lambda u: u.is_superuser)
@login_required
def restore_records(request):
    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        if csv_file:
            try:
                # Decode the file content
                decoded_file = csv_file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                
                for row in reader:
                    # Handle user fields
                    if 'created_by' in row and row['created_by']:
                        try:
                            row['created_by'] = User.objects.get(username=row['created_by'])
                        except User.DoesNotExist:
                            row['created_by'] = None
                    
                    if 'edited_by' in row and row['edited_by']:
                        try:
                            row['edited_by'] = User.objects.get(username=row['edited_by'])
                        except User.DoesNotExist:
                            row['edited_by'] = None

                    # Handle other fields as before
                    if row.get('created_at'):
                        try:
                            created_at = parser.parse(row['created_at'])
                            if timezone.is_naive(created_at):
                                created_at = timezone.make_aware(created_at)  # Make it timezone-aware
                            row['created_at'] = created_at
                        except (ValueError, TypeError):
                            row['created_at'] = timezone.now()  # Default to now if parsing fails
                    
                    if row.get('updated_at'):
                        try:
                            updated_at = parser.parse(row['updated_at'])
                            if timezone.is_naive(updated_at):
                                updated_at = timezone.make_aware(updated_at)  # Make it timezone-aware
                            row['updated_at'] = updated_at
                        except (ValueError, TypeError):
                            row['updated_at'] = timezone.now()  # Default to now if parsing fails
                    
                    if row.get('delivery_date'):
                        try:
                            delivery_date = parser.parse(row['delivery_date']).date()
                            row['delivery_date'] = delivery_date
                        except (ValueError, TypeError) as e:
                            raise ValueError(f"Invalid delivery_date format: {row['delivery_date']}")

                    # Ensure delivery_date is not None
                    if 'delivery_date' not in row or row['delivery_date'] is None:
                        raise ValueError("Delivery date is required and cannot be None.")

                    # Convert string time to time object
                    if row.get('delivery_time'):
                        try:
                            delivery_time = parser.parse(row['delivery_time']).time()
                            row['delivery_time'] = delivery_time
                        except (ValueError, TypeError) as e:
                            raise ValueError(f"Invalid delivery_time format: {row['delivery_time']}")

                    # Handle boolean fields
                    boolean_fields = [
                        'alive', 'vitamin_k', 'ttc_eye_ointment', 'bcg_given', 
                        'opv_given', 'hiv_test_accepted', 'hiv_retesting_accepted',
                        'known_hiv_positive', 'counseled_on_feeding_options',
                        'live_birth_died_before_arrival', 
                        'live_birth_died_after_arrival_or_delivery'
                    ]
                    
                    for field in boolean_fields:
                        if field in row:
                            if isinstance(row[field], str):
                                row[field] = str(row[field]).lower() in ('true', '1', 't', 'yes')
                            elif row[field] in ('', None):
                                row[field] = None

                    # Handle numeric fields
                    numeric_fields = ['age', 'weight_in_grams', 'mrn_newborn', 'stillbirth']
                    for field in numeric_fields:
                        if field in row and row[field]:
                            try:
                                row[field] = int(float(row[field]))
                            except (ValueError, TypeError):
                                row[field] = None

                    # Remove any empty string values for nullable fields
                    for key, value in row.items():
                        if value == '':
                            row[key] = None

                    # Handle foreign key fields
                    if 'created_by_id' in row:
                        row['created_by_id'] = row.pop('created_by')
                    if 'edited_by_id' in row:
                        row['edited_by_id'] = row.pop('edited_by')

                    # Create the record
                    try:
                        DeliveryRecord.objects.create(**row)
                    except Exception as e:
                        raise ValueError(f"Error creating record: {str(e)}, Data: {row}")

                messages.success(request, 'Data restored successfully from CSV')
                return redirect('core:record_list')
                
            except (ValueError, KeyError) as e:
                messages.error(request, f'Error processing CSV file: {str(e)}')
            except ValidationError as e:
                messages.error(request, f'Validation error: {str(e)}')
            except Exception as e:
                messages.error(request, f'Unexpected error: {str(e)}')
                
    return render(request, 'core/restore_page.html', {'form': RestoreForm()})

@user_passes_test(lambda u: u.is_superuser)
def user_productivity_details(request, user_id):
    user = get_object_or_404(User, id=user_id)
    
    # Get date range from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # If no dates provided, use last 30 days
    today = timezone.now().date()
    if not start_date or not end_date:
        end_date = today
        start_date = end_date - timedelta(days=30)
    else:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            end_date = today
            start_date = end_date - timedelta(days=30)
    
    # Get user's records
    records = DeliveryRecord.objects.filter(
        created_by=user,
        created_at__date__range=[start_date, end_date]
    ).order_by('-created_at')
    
    # Calculate statistics
    daily_records = records.filter(created_at__date=today).count()
    weekly_records = records.filter(
        created_at__date__range=[today - timedelta(days=7), today]
    ).count()
    monthly_records = records.filter(
        created_at__date__range=[today.replace(day=1), today]
    ).count()
    
    context = {
        'user_detail': user,
        'records': records,
        'daily_records': daily_records,
        'weekly_records': weekly_records,
        'monthly_records': monthly_records,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'core/user_productivity_details.html', context)

@user_passes_test(lambda u: u.is_superuser)
def recording_trends(request):
    range_type = request.GET.get('range', 'month')
    now = timezone.now()
    today = now.date()

    # Set date range based on selected range
    if range_type == 'day':
        # Last 24 hours
        start_datetime = now - timedelta(hours=24)
        end_datetime = now
        date_format = '%H:00'  # Show only hour
        records = DeliveryRecord.objects.filter(
            created_at__range=[start_datetime, end_datetime]
        ).annotate(
            date=TruncHour('created_at')
        ).values('date').annotate(
            count=Count('serial_number')
        ).order_by('date')
        
        # Generate all hours
        current = start_datetime
        labels = []
        values = []
        while current <= end_datetime:
            labels.append(current.strftime(date_format))
            record = next((r for r in records if r['date'] == current), None)
            values.append(record['count'] if record else 0)
            current += timedelta(hours=1)

    elif range_type == 'week':
        # Last 7 days
        start_date = today - timedelta(days=6)
        end_date = today
        date_format = '%a'  # Show day name (Mon, Tue, etc.)
        records = DeliveryRecord.objects.filter(
            created_at__date__range=[start_date, end_date]
        ).annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(
            count=Count('serial_number')
        ).order_by('date')
        
        # Generate all days
        labels = []
        values = []
        current = start_date
        while current <= end_date:
            labels.append(current.strftime(date_format))
            record = next((r for r in records if r['date'] == current), None)
            values.append(record['count'] if record else 0)
            current += timedelta(days=1)

    elif range_type == 'month':
        # Last 30 days
        start_date = today - timedelta(days=29)
        end_date = today
        date_format = '%b %d'  # Show month and day (Jan 01)
        records = DeliveryRecord.objects.filter(
            created_at__date__range=[start_date, end_date]
        ).annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(
            count=Count('serial_number')
        ).order_by('date')
        
        # Generate all days
        labels = []
        values = []
        current = start_date
        while current <= end_date:
            labels.append(current.strftime(date_format))
            record = next((r for r in records if r['date'] == current), None)
            values.append(record['count'] if record else 0)
            current += timedelta(days=1)

    elif range_type == 'total':
        # All time by month
        first_record = DeliveryRecord.objects.earliest('created_at')
        start_date = first_record.created_at.date() if first_record else today
        start_date = start_date.replace(day=1)  # Start from first day of month
        end_date = today
        date_format = '%b %Y'  # Show month and year (Jan 2024)
        records = DeliveryRecord.objects.filter(
            created_at__date__range=[start_date, end_date]
        ).annotate(
            date=TruncMonth('created_at')
        ).values('date').annotate(
            count=Count('serial_number')
        ).order_by('date')
        
        # Generate all months
        labels = []
        values = []
        current = start_date
        while current <= end_date:
            labels.append(current.strftime(date_format))
            record = next((r for r in records if r['date'].date().replace(day=1) == current), None)
            values.append(record['count'] if record else 0)
            current = (current.replace(day=1) + timedelta(days=32)).replace(day=1)

    elif range_type == 'custom':
        try:
            start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
            date_format = '%Y-%m-%d'
            records = DeliveryRecord.objects.filter(
                created_at__date__range=[start_date, end_date]
            ).annotate(
                date=TruncDate('created_at')
            ).values('date').annotate(
                count=Count('serial_number')
            ).order_by('date')
            
            # Generate all days in custom range
            labels = []
            values = []
            current = start_date
            while current <= end_date:
                labels.append(current.strftime(date_format))
                record = next((r for r in records if r['date'] == current), None)
                values.append(record['count'] if record else 0)
                current += timedelta(days=1)
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid date format'}, status=400)
    else:
        return JsonResponse({'error': 'Invalid range type'}, status=400)

    return JsonResponse({
        'labels': labels,
        'values': values
    })
